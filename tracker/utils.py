import csv
import openpyxl 
from PIL import Image, ImageDraw, ImageFont, ExifTags, ImageOps
from io import BytesIO
from django.core.files.base import ContentFile
import datetime

# ==========================================
# 1. GPS & WATERMARK LOGIC
# ==========================================
def _convert_to_degrees(value):
    d = float(value[0])
    m = float(value[1])
    s = float(value[2])
    return d + (m / 60.0) + (s / 3600.0)

def get_gps_from_image(image_field):
    try:
        img = Image.open(image_field)
        exif_data = img._getexif()
        if not exif_data: return None, None
        
        gps_info = exif_data.get(34853) # 34853 is GPSInfo
        if not gps_info: return None, None

        lat_gps = gps_info.get(2)
        lat_ref = gps_info.get(1)
        lon_gps = gps_info.get(4)
        lon_ref = gps_info.get(3)

        if lat_gps and lat_ref and lon_gps and lon_ref:
            lat = _convert_to_degrees(lat_gps)
            if lat_ref != 'N': lat = -lat
            lon = _convert_to_degrees(lon_gps)
            if lon_ref != 'E': lon = -lon
            return f"{lat:.6f}", f"{lon:.6f}"
    except Exception:
        pass
    return None, None

def watermark_image(image_field, lat, lon):
    try:
        img = Image.open(image_field)
        img = ImageOps.exif_transpose(img)
        img = img.convert('RGB')
        draw = ImageDraw.Draw(img)
        width, height = img.size
        font_size = int(width * 0.04)
        
        try: font = ImageFont.truetype("arial.ttf", font_size)
        except: font = ImageFont.load_default()

        draw.rectangle([(0, 0), (width, font_size * 2)], fill=(0, 0, 0, 160))
        draw.text((20, 20), "Nexsafe", fill=(255, 165, 0), font=font)
        
        bottom = font_size * 2.5
        draw.rectangle([(0, height - bottom), (width, height)], fill=(0, 0, 0, 160))
        draw.text((20, height - bottom + 10), f"{lat}, {lon}", fill=(255, 255, 255), font=font)
        
        buffer = BytesIO()
        img.save(buffer, format='JPEG', quality=85)
        return ContentFile(buffer.getvalue())
    except Exception:
        image_field.seek(0)
        return image_field

# ==========================================
# 2. EXCEL/CSV HELPERS (MISSING IN YOUR CODE)
# ==========================================

def get_file_headers(file_field):
    """Returns a list of column headers from the file."""
    if not file_field:
        return []

    try:
        # Ensure file is open and at the start
        try: file_field.open('rb')
        except: pass
        file_field.seek(0)
        
        filename = file_field.name.lower()
        headers = []

        if filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_field, data_only=True)
            sheet = workbook.active
            # Get first row only
            headers = [str(cell.value).strip() for cell in sheet[1] if cell.value]
        else:
            decoded_file = file_field.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(decoded_file)
            headers = next(reader, [])
            headers = [h.strip() for h in headers if h]
            
        return headers
    except Exception as e:
        print(f"Header Read Error: {e}")
        return []

def get_dropdown_options(file_field, column_name):
    """Reads file and returns unique values from a specific column."""
    if not file_field: return []
    
    options = set()
    try:
        try: file_field.open('rb')
        except: pass
        file_field.seek(0)
        filename = file_field.name.lower()
        
        if filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_field, data_only=True)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            
            col_index = None
            for idx, header in enumerate(headers):
                if header == column_name.strip():
                    col_index = idx
                    break
            
            if col_index is not None:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    val = row[col_index]
                    if val: options.add(str(val).strip())

        else:
            decoded_file = file_field.read().decode('utf-8-sig').splitlines()
            reader = csv.DictReader(decoded_file)
            reader.fieldnames = [name.strip() for name in reader.fieldnames]
            
            if column_name.strip() in reader.fieldnames:
                for row in reader:
                    val = row.get(column_name.strip())
                    if val: options.add(val.strip())

        sorted_options = sorted(list(options))
        return [(opt, opt) for opt in sorted_options]
        
    except Exception:
        return []